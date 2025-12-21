from helpers import Helpers

import re
from openpyxl.utils import get_column_letter

class SimpleFormule:
    def __init__(self, module: str, formule: str, is_simple: bool):
        self.module = module
        self.formule = formule[1:]
        self.is_simple = is_simple

    def exec(self):
        return self.resolve_expression(self.formule)

    def excel_if(self, if_str: str):
        args = Helpers.split_excel_args(if_str)
        if len(args) < 2:
            raise ValueError(f'A excel if statement need at least 2 section a condition, a value if contion is true')

        condition = self.resolve_expression(args[0])
        if_content = self.resolve_expression(args[1])
        else_content = self.resolve_expression(args[2]) if len(args) == 3 else 'nil'

        return (
            f'if {condition}\n'
            f'    {if_content}\n'
            f'else\n'
            f'    {else_content}\n'
            f'end\n'
        )

    def excel_and_or(self, exp_str: str, is_or = False):
        res = []

        for condition in Helpers.split_excel_args(exp_str):
            part1, comp, part2 = Helpers.separate_condition(condition)

            part1 = self.resolve_expression(part1)
            part2 = self.resolve_expression(part2)

            res.append(f'{part1} {comp} {part2}')

        if is_or:
            return f"({' || '.join(res)})"
        else:
            return f"({' && '.join(res)})"

    def excel_not(self, not_str: str):
        try:
            part1, comp, part2 = Helpers.separate_condition(not_str)

            part1 = self.resolve_expression(part1)
            part2 = self.resolve_expression(part2)

            return (f'!({part1} {comp} {part2})')
        except Exception:
            return f'!{self.resolve_expression(not_str)}'

    def excel_sum(self, module: str, exp: str):
        terms = []

        for current_range in exp.split(','):
            cells = current_range.strip().split(':')
            
            end_cell_name = re.sub(r'\W+', '', cells[-1].lower())

            start_reference_sheetname = Helpers.reference_sheetname(module, cells[0])
            end_reference_sheetname = start_reference_sheetname.split('(:')[0] + f"(:{end_cell_name})"

            start_cell = Helpers.cell_to_int(start_reference_sheetname.split('(:')[1][:-1])
            end_cell = Helpers.cell_to_int(end_reference_sheetname.split('(:')[1][:-1])

            end_sheet_reference = cells[0].index('!') - 1
            start_sheet_reference = end_sheet_reference - Helpers.end_single_quote(cells[0][:end_sheet_reference][::-1], 0)
            sheetname_referenced = re.sub(r'\W+', '', cells[0][start_sheet_reference:end_sheet_reference]).title().replace(' ', '')

            for col in range(start_cell[0], end_cell[0] + 1, 1):
                for row in range(start_cell[1], end_cell[1] + 1, 1):
                    terms.append(f'{module}::{sheetname_referenced}.cell(:{get_column_letter(col + 1)}{row})')

        return f'{terms}.sum'

    def resolve_expression(self, expression: str) -> str:
        expression = expression.strip()
        expression = expression.replace('_xlfn._xlws', '')

        if expression.startswith('IF('):
            return self.excel_if(expression[3:-1])
        elif expression.startswith('NOT('):
            return self.excel_not(expression[4:-1])
        elif expression.startswith('AND('):
            return self.excel_and_or(expression[4:-1], False)
        elif expression.startswith('OR('):
            return self.excel_and_or(expression[3:-1], True)
        elif expression.startswith('SUM('):
            return self.excel_sum(self.module, expression[3:-1])
        elif '!' in expression:
            return Helpers.reference_sheetname(self.module, expression)
        elif Helpers.is_cell_reference(expression):
            return expression.lower()
        elif '&' in expression:
            return ' + '.join([self.resolve_expression(p.strip()) + '.to_s' for p in expression.split('&')])
        elif expression.startswith('FALSE('):
            return 'false'
        else:
            return expression
