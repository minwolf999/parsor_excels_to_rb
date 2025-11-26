import re
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

from helpers import Helpers
from simple_formule import SimpleFormule
from array_formule import ArrayFormule

class ExcelToRb:
    def __init__(self, workbook_path: str, wb: Workbook, excel_name: str):
        self.workbook_path = workbook_path
        self.wb = wb
        self.excel_name = ''.join(word.capitalize() for word in excel_name.split('_'))

    def exec(self):
        for index, sheetname in enumerate(self.wb.sheetnames):
            clean_sheetname = re.sub(r'\W+', '', sheetname).title().replace(' ', '')

            with open(f'{self.workbook_path}/{clean_sheetname}.rb', 'w') as f:
                f.write(self.excel_to_rb(sheetname, clean_sheetname))

            if index == 2:
                exit(0)

    def excel_to_rb(self, sheetname: str, clean_sheetname: str):
        sheet = self.wb[sheetname]

        content_file = '# frozen_string_literal: true\n\n'
        content_file += f'module {self.excel_name}\n'
        content_file += f'  class {clean_sheetname}\n'
        init = '    def initialize\n'
        functions = ''

        for row in range(1, sheet.max_row + 1):
            for column in range(1, sheet.max_column + 1):
                cell = sheet.cell(row, column)

                cell_value = cell.value
                function_args = ''

                if cell_value is None:
                    init += f"      {get_column_letter(column).lower()}{row} = nil\n"
                    continue
                elif isinstance(cell_value, str) and cell_value.startswith("="):
                    cell_value = SimpleFormule(cell_value, True).exec()
                elif isinstance(cell_value, ArrayFormula):
                    print()
                    print(cell_value.text)
                    print(cell_value)
                    print()

                    cell_value = ArrayFormule(cell_value.text, False).exec()
                elif not Helpers.is_number(cell_value):
                    cell_value = json.dumps(cell_value, ensure_ascii=False)

                functions += f'    def {get_column_letter(column).lower()}{row}({function_args})\n'
                functions += f'      {cell_value}\n'
                functions += f'    end\n'

        init += '    end\n'
        content_file += f'{init}\n{functions}'
        content_file += '  end\n'
        return content_file.rstrip() + '\nend\n'
